"""
Audit View Widget for Admin Application
Displays audit logs and provides export functionality
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QTableWidget,
    QTableWidgetItem, QMessageBox, QGroupBox,
    QFileDialog, QHeaderView, QAbstractItemView,
    QDateEdit, QComboBox
)
from PySide6.QtCore import Qt, QDate

from PySide6.QtGui import QKeySequence, QShortcut

from core.database import get_session, close_session
from core.models import AuditLog, User, Payment, IsoMessage
from shared_ui.components import attach_empty_state
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
import os


class AuditViewWidget(QWidget):
    """Widget for viewing audit logs and exporting data"""
    
    def __init__(self):
        super().__init__()
        self._audit_offset = 0
        self.init_ui()
        self.load_audit_logs()
    
    def init_ui(self):
        """Initialize the user interface"""
        layout = QVBoxLayout(self)
        layout.setSpacing(20)
        
        # Header
        header = QLabel("Auditoría y Exportación de Datos")
        header.setProperty("class", "subheader")
        layout.addWidget(header)
        
        # Filters
        filter_group = self.create_filters()
        layout.addWidget(filter_group)
        
        # Audit log table
        log_group = self.create_audit_table()
        layout.addWidget(log_group)
        
        # Export section
        export_group = self.create_export_section()
        layout.addWidget(export_group)

        QShortcut(QKeySequence("F5"), self).activated.connect(self.load_audit_logs)
        QShortcut(QKeySequence("Ctrl+E"), self).activated.connect(self.export_audit_to_excel)
    
    def create_filters(self) -> QGroupBox:
        """Create filter controls"""
        group = QGroupBox("Filtros")
        layout = QHBoxLayout()
        
        # Date from
        layout.addWidget(QLabel("Desde:"))
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(QDate.currentDate().addMonths(-1))
        self.date_from.setDisplayFormat("dd/MM/yyyy")
        layout.addWidget(self.date_from)
        
        # Date to
        layout.addWidget(QLabel("Hasta:"))
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QDate.currentDate())
        self.date_to.setDisplayFormat("dd/MM/yyyy")
        layout.addWidget(self.date_to)
        
        layout.addStretch()
        
        # Apply filter button
        apply_btn = QPushButton("Aplicar Filtros")
        apply_btn.clicked.connect(self.load_audit_logs)
        layout.addWidget(apply_btn)
        
        group.setLayout(layout)
        return group
    
    def create_audit_table(self) -> QGroupBox:
        """Create the audit log table"""
        group = QGroupBox("Registro de Auditoría")
        layout = QVBoxLayout()
        
        # Table
        self.audit_table = QTableWidget()
        self.audit_table.setColumnCount(4)
        self.audit_table.setHorizontalHeaderLabels([
            "Fecha/Hora", "Usuario", "Acción", "Detalles"
        ])
        
        # Table settings
        self.audit_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.audit_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.audit_table.horizontalHeader().setStretchLastSection(True)
        self.audit_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        
        self.audit_table.setAlternatingRowColors(True)
        layout.addWidget(self.audit_table)
        attach_empty_state(self.audit_table, "No hay registros para el rango seleccionado")

        self.audit_count_label = QLabel("Cargando...")
        self.audit_count_label.setStyleSheet("font-size: 9pt; color: #605E5C;")
        layout.addWidget(self.audit_count_label, alignment=Qt.AlignLeft)

        self.audit_load_more_btn = QPushButton("Cargar más...")
        self.audit_load_more_btn.setProperty("class", "secondary")
        self.audit_load_more_btn.setVisible(False)
        self.audit_load_more_btn.clicked.connect(self._load_more_audit)
        layout.addWidget(self.audit_load_more_btn, alignment=Qt.AlignLeft)

        # Refresh button
        refresh_btn = QPushButton("🔄 Actualizar")
        refresh_btn.setProperty("class", "secondary")
        refresh_btn.clicked.connect(self.load_audit_logs)
        layout.addWidget(refresh_btn, alignment=Qt.AlignRight)
        
        group.setLayout(layout)
        return group
    
    def create_export_section(self) -> QGroupBox:
        """Create export controls"""
        group = QGroupBox("Exportar Datos")
        layout = QHBoxLayout()
        
        # Export buttons
        export_audit_btn = QPushButton("📊 Exportar Auditoría a Excel")
        export_audit_btn.clicked.connect(self.export_audit_to_excel)
        layout.addWidget(export_audit_btn)
        
        export_payments_btn = QPushButton("💰 Exportar Pagos a Excel")
        export_payments_btn.clicked.connect(self.export_payments_to_excel)
        layout.addWidget(export_payments_btn)
        
        export_iso_btn = QPushButton("📄 Exportar Mensajes ISO 20022")
        export_iso_btn.clicked.connect(self.export_iso_messages)
        layout.addWidget(export_iso_btn)

        cierre_btn = QPushButton("📥 Cierre de Día (camt.053)")
        cierre_btn.setProperty("class", "success")
        cierre_btn.setToolTip(
            "Genera el estado de cuenta del día (camt.053)\n"
            "con todos los pagos completados y simulados de hoy.\n"
            "Equivalente bancario: Bank Statement."
        )
        cierre_btn.clicked.connect(self.generate_cierre_dia)
        layout.addWidget(cierre_btn)

        layout.addStretch()
        
        group.setLayout(layout)
        return group
    
    def load_audit_logs(self):
        """Load audit logs into the table"""
        try:
            session = get_session()
            
            # Get date range
            date_from = self.date_from.date().toPython()
            date_to = self.date_to.date().toPython()
            
            # Convert to datetime
            datetime_from = datetime.combine(date_from, datetime.min.time())
            datetime_to = datetime.combine(date_to, datetime.max.time())
            
            # Query logs
            self._audit_offset = 0
            total_count = session.query(AuditLog).filter(
                AuditLog.timestamp >= datetime_from,
                AuditLog.timestamp <= datetime_to
            ).count()
            logs = session.query(AuditLog).filter(
                AuditLog.timestamp >= datetime_from,
                AuditLog.timestamp <= datetime_to
            ).order_by(AuditLog.timestamp.desc()).limit(200).all()

            self.audit_table.setRowCount(len(logs))

            for row, log in enumerate(logs):
                # Timestamp
                timestamp_str = log.timestamp.strftime("%d/%m/%Y %H:%M:%S")
                self.audit_table.setItem(row, 0, QTableWidgetItem(timestamp_str))

                # User
                user_name = log.user.full_name if log.user else "Sistema"
                self.audit_table.setItem(row, 1, QTableWidgetItem(user_name))

                # Action
                self.audit_table.setItem(row, 2, QTableWidgetItem(log.action))

                # Details
                details = log.details or "—"
                self.audit_table.setItem(row, 3, QTableWidgetItem(details))

            self.audit_count_label.setText(f"Mostrando {len(logs)} de {total_count} registros")
            self.audit_load_more_btn.setVisible(total_count > 200)
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Error al cargar registros de auditoría:\n{str(e)}"
            )
        finally:
            close_session()

    def _load_more_audit(self):
        """Append next 200 audit records"""
        try:
            session = get_session()
            self._audit_offset += 200

            date_from = self.date_from.date().toPython()
            date_to = self.date_to.date().toPython()
            datetime_from = datetime.combine(date_from, datetime.min.time())
            datetime_to = datetime.combine(date_to, datetime.max.time())

            more_logs = session.query(AuditLog).filter(
                AuditLog.timestamp >= datetime_from,
                AuditLog.timestamp <= datetime_to
            ).order_by(AuditLog.timestamp.desc()).offset(self._audit_offset).limit(200).all()

            current_rows = self.audit_table.rowCount()
            self.audit_table.setRowCount(current_rows + len(more_logs))

            total_count = session.query(AuditLog).filter(
                AuditLog.timestamp >= datetime_from,
                AuditLog.timestamp <= datetime_to
            ).count()

            for row_idx, log in enumerate(more_logs):
                row = current_rows + row_idx
                self.audit_table.setItem(row, 0, QTableWidgetItem(
                    log.timestamp.strftime("%d/%m/%Y %H:%M:%S")
                ))
                self.audit_table.setItem(row, 1, QTableWidgetItem(
                    log.user.full_name if log.user else "Sistema"
                ))
                self.audit_table.setItem(row, 2, QTableWidgetItem(log.action))
                self.audit_table.setItem(row, 3, QTableWidgetItem(log.details or "—"))

            loaded = self.audit_table.rowCount()
            self.audit_count_label.setText(f"Mostrando {loaded} de {total_count} registros")
            self.audit_load_more_btn.setVisible(loaded < total_count)

        except Exception as e:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.critical(self, "Error", f"Error al cargar más registros:\n{str(e)}")
        finally:
            close_session()

    def export_audit_to_excel(self):
        """Export audit logs to Excel"""
        try:
            # Get save location
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar Auditoría",
                f"auditoria_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # Get data
            session = get_session()
            logs = session.query(AuditLog).order_by(AuditLog.timestamp.desc()).all()
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Auditoría"
            
            # Headers
            headers = ["Fecha/Hora", "Usuario", "Acción", "Detalles"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")

            # Data
            for row, log in enumerate(logs, 2):
                ws.cell(row=row, column=1, value=log.timestamp.strftime("%d/%m/%Y %H:%M:%S"))
                ws.cell(row=row, column=2, value=log.user.full_name if log.user else "Sistema")
                ws.cell(row=row, column=3, value=log.action)
                ws.cell(row=row, column=4, value=log.details or "")
            
            # Save
            wb.save(file_path)
            from shared_ui.components import Toast
            Toast.show_message(self, f"✓ Auditoría exportada")
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Error al exportar auditoría:\n{str(e)}"
            )
        finally:
            close_session()
    
    def export_payments_to_excel(self):
        """Export payments to Excel"""
        try:
            # Get save location
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar Pagos",
                f"pagos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # Get data
            session = get_session()
            payments = session.query(Payment).order_by(Payment.timestamp.desc()).all()
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Pagos"
            
            # Headers
            headers = [
                "UETR", "Hash XRPL", "Fecha/Hora", "Productor",
                "Operador", "Monto", "Moneda", "Monto MXN", "Estado"
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
            
            # Data
            for row, payment in enumerate(payments, 2):
                ws.cell(row=row, column=1, value=payment.uetr)
                ws.cell(row=row, column=2, value=payment.xrpl_tx_hash)
                ws.cell(row=row, column=3, value=payment.timestamp.strftime("%d/%m/%Y %H:%M:%S"))
                ws.cell(row=row, column=4, value=payment.producer.name)
                ws.cell(row=row, column=5, value=payment.operator.full_name)
                ws.cell(row=row, column=6, value=payment.amount)
                ws.cell(row=row, column=7, value=payment.currency)
                ws.cell(row=row, column=8, value=payment.amount_mxn or 0)
                ws.cell(row=row, column=9, value=payment.status.value)
            
            # Save
            wb.save(file_path)
            from shared_ui.components import Toast
            Toast.show_message(self, "✓ Pagos exportados")
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Error al exportar pagos:\n{str(e)}"
            )
        finally:
            close_session()
    
    def export_iso_messages(self):
        """Export ISO 20022 messages to XML files"""
        try:
            # Get save directory
            dir_path = QFileDialog.getExistingDirectory(
                self,
                "Seleccionar Carpeta para Mensajes ISO"
            )
            
            if not dir_path:
                return
            
            # Get data
            session = get_session()
            messages = session.query(IsoMessage).all()
            
            if not messages:
                QMessageBox.information(
                    self,
                    "Sin Datos",
                    "No hay mensajes ISO 20022 para exportar."
                )
                return
            
            # Export each message
            for msg in messages:
                uetr = msg.payment.uetr if msg.payment else "no-payment"
                filename = f"{msg.message_type.value}_{uetr}.xml"
                file_path = os.path.join(dir_path, filename)
                
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(msg.xml_content)
            
            from shared_ui.components import Toast
            Toast.show_message(self, f"✓ {len(messages)} mensajes ISO exportados")
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Error al exportar mensajes ISO:\n{str(e)}"
            )
        finally:
            close_session()

    def generate_cierre_dia(self):
        """Generate end-of-day camt.053 bank statement for today's payments"""
        from datetime import date, timezone
        from core.iso_generator import ISO20022Generator
        from core.models import Payment, IsoMessage, MessageType, PaymentStatus
        from core.xrpl_client import XRPLClient
        import uuid

        try:
            session = get_session()

            # All completed/simulated payments for today
            today = datetime.combine(date.today(), datetime.min.time())
            tomorrow = datetime.combine(date.today(), datetime.max.time())

            today_payments = session.query(Payment).filter(
                Payment.timestamp >= today,
                Payment.timestamp <= tomorrow,
                Payment.status.in_([PaymentStatus.COMPLETED, PaymentStatus.SIMULATED])
            ).order_by(Payment.timestamp.asc()).all()

            if not today_payments:
                QMessageBox.information(
                    self,
                    "Sin Pagos",
                    "No hay pagos completados hoy para generar el estado de cuenta."
                )
                return

            # Try to get current XRP balance (best-effort)
            opening_balance = 0.0
            try:
                # Use the first payment's operator wallet address
                first_operator = today_payments[0].operator
                client = XRPLClient()
                bal = client.get_balance(first_operator.xrpl_address)
                # opening balance = current balance (simplified: no historical reconstruction)
                opening_balance = bal.get("xrp", 0.0)
            except Exception:
                opening_balance = 0.0

            statement_data = {
                "statement_id": f"STMT-{date.today().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}",
                "account_id": today_payments[0].operator.xrpl_address,
                "account_name": today_payments[0].operator.full_name,
                "opening_balance": opening_balance,
                "from_date": today,
                "to_date": tomorrow,
            }

            gen = ISO20022Generator()
            xml_str = gen.generate_camt053(today_payments, statement_data)

            # Save to DB (payment_id=None — statement covers multiple payments)
            iso_msg = IsoMessage(
                payment_id=None,
                message_type=MessageType.CAMT_053,
                xml_content=xml_str,
            )
            session.add(iso_msg)
            session.commit()

            # Offer to save file
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar Cierre de Día",
                f"camt053_cierre_{date.today().strftime('%Y%m%d')}.xml",
                "XML Files (*.xml)"
            )

            if file_path:
                with open(file_path, "w", encoding="utf-8") as f:
                    f.write(xml_str)
                QMessageBox.information(
                    self,
                    "Cierre Generado",
                    f"Estado de cuenta camt.053 generado con {len(today_payments)} entradas.\n"
                    f"Guardado en: {file_path}"
                )
            else:
                QMessageBox.information(
                    self,
                    "Cierre Generado",
                    f"Estado de cuenta camt.053 guardado en base de datos "
                    f"({len(today_payments)} entradas). Sin exportar a archivo."
                )

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al generar cierre de día:\n{str(e)}")
        finally:
            close_session()
