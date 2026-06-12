"""
History View Widget
Display payment history and export functionality
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QTableWidget,
    QTableWidgetItem, QMessageBox, QGroupBox,
    QFileDialog, QHeaderView, QAbstractItemView,
    QDateEdit, QComboBox
)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QColor

from core.database import get_session, close_session
from core.models import Payment, User
from core.utils import format_currency, format_datetime_display
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill


class HistoryViewWidget(QWidget):
    """Widget for viewing payment history"""
    
    def __init__(self, operator: User):
        super().__init__()
        self.operator = operator
        self._history_offset = 0
        self.init_ui()
        self.load_history()
    
    def init_ui(self):
        """Initialize the user interface"""
        layout = QVBoxLayout(self)
        layout.setSpacing(20)
        
        # Header
        header_layout = QHBoxLayout()
        
        title = QLabel("Historial de Pagos")
        title.setProperty("class", "subheader")
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        # Refresh button
        refresh_btn = QPushButton("🔄 Actualizar")
        refresh_btn.setProperty("class", "secondary")
        refresh_btn.clicked.connect(self.load_history)
        header_layout.addWidget(refresh_btn)
        
        # Export button
        export_btn = QPushButton("📊 Exportar a Excel")
        export_btn.clicked.connect(self.export_to_excel)
        header_layout.addWidget(export_btn)
        
        layout.addLayout(header_layout)

        # Filters
        self.filter_group = self.create_filters()
        layout.addWidget(self.filter_group)

        # Payment table
        self.payment_table = QTableWidget()
        self.payment_table.setColumnCount(7)
        self.payment_table.setHorizontalHeaderLabels([
            "Fecha/Hora", "Productor", "Peso (kg)",
            "Precio/kg", "Total MXN", "Token", "Estado"
        ])
        
        # Table settings
        self.payment_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.payment_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.payment_table.horizontalHeader().setStretchLastSection(True)
        self.payment_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.payment_table.doubleClicked.connect(self.show_payment_details)
        
        layout.addWidget(self.payment_table)

        self.load_more_btn = QPushButton("Cargar más...")
        self.load_more_btn.setProperty("class", "secondary")
        self.load_more_btn.setVisible(False)
        self.load_more_btn.clicked.connect(self._load_more)
        layout.addWidget(self.load_more_btn)

        # Summary
        self.summary_label = QLabel()
        self.summary_label.setStyleSheet("font-size: 10pt; color: #605E5C; padding: 10px;")
        layout.addWidget(self.summary_label)
    
    def create_filters(self) -> QGroupBox:
        """Create date/producer/status filter bar"""
        group = QGroupBox("Filtros")
        layout = QHBoxLayout()

        layout.addWidget(QLabel("Desde:"))
        self.filter_date_from = QDateEdit()
        self.filter_date_from.setCalendarPopup(True)
        self.filter_date_from.setDate(QDate.currentDate().addMonths(-1))
        self.filter_date_from.setDisplayFormat("dd/MM/yyyy")
        layout.addWidget(self.filter_date_from)

        layout.addWidget(QLabel("Hasta:"))
        self.filter_date_to = QDateEdit()
        self.filter_date_to.setCalendarPopup(True)
        self.filter_date_to.setDate(QDate.currentDate())
        self.filter_date_to.setDisplayFormat("dd/MM/yyyy")
        layout.addWidget(self.filter_date_to)

        layout.addWidget(QLabel("Estado:"))
        self.filter_status = QComboBox()
        self.filter_status.addItems(["Todos", "Completado", "Simulado", "Pendiente", "Fallido"])
        layout.addWidget(self.filter_status)

        layout.addStretch()

        apply_btn = QPushButton("Aplicar")
        apply_btn.clicked.connect(self.load_history)
        layout.addWidget(apply_btn)

        group.setLayout(layout)
        return group

    def _build_filtered_query(self, session):
        """Build a Payment query applying the active date/status filters."""
        from datetime import datetime as dt
        from core.models import PaymentStatus

        query = session.query(Payment).filter_by(operator_id=self.operator.id)

        # Date filters
        date_from = self.filter_date_from.date().toPython()
        date_to = self.filter_date_to.date().toPython()
        datetime_from = dt.combine(date_from, dt.min.time())
        datetime_to = dt.combine(date_to, dt.max.time())
        query = query.filter(
            Payment.timestamp >= datetime_from,
            Payment.timestamp <= datetime_to
        )

        # Status filter
        status_text = self.filter_status.currentText()
        status_map = {
            "Completado": PaymentStatus.COMPLETED,
            "Simulado": PaymentStatus.SIMULATED,
            "Pendiente": PaymentStatus.PENDING,
            "Fallido": PaymentStatus.FAILED,
        }
        if status_text in status_map:
            query = query.filter(Payment.status == status_map[status_text])

        return query

    def load_history(self):
        """Load payment history"""
        try:
            self._history_offset = 0
            session = get_session()

            # Get payments for this operator, applying active filters with pagination
            payments = self._build_filtered_query(session).order_by(
                Payment.timestamp.desc()
            ).offset(self._history_offset).limit(200).all()

            total_count = self._build_filtered_query(session).count()

            self.payment_table.setRowCount(len(payments))

            total_mxn = 0.0
            total_kg = 0.0

            for row, payment in enumerate(payments):
                # Timestamp
                timestamp_str = format_datetime_display(payment.timestamp)
                item = QTableWidgetItem(timestamp_str)
                item.setData(Qt.UserRole, payment.id)
                self.payment_table.setItem(row, 0, item)

                # Producer
                self.payment_table.setItem(row, 1, QTableWidgetItem(payment.producer.name))

                # Weight
                if payment.delivery:
                    weight_str = f"{float(payment.delivery.weight_kg):.2f}"
                    self.payment_table.setItem(row, 2, QTableWidgetItem(weight_str))
                    total_kg += float(payment.delivery.weight_kg)

                    # Price per kg
                    price_str = format_currency(float(payment.delivery.price_per_kg), "MXN")
                    self.payment_table.setItem(row, 3, QTableWidgetItem(price_str))
                else:
                    self.payment_table.setItem(row, 2, QTableWidgetItem("—"))
                    self.payment_table.setItem(row, 3, QTableWidgetItem("—"))

                # Total MXN
                if payment.amount_mxn:
                    total_str = format_currency(float(payment.amount_mxn), "MXN")
                    self.payment_table.setItem(row, 4, QTableWidgetItem(total_str))
                    total_mxn += float(payment.amount_mxn)
                else:
                    self.payment_table.setItem(row, 4, QTableWidgetItem("—"))

                # Token
                token_str = f"{float(payment.amount):.6f} {payment.currency}"
                self.payment_table.setItem(row, 5, QTableWidgetItem(token_str))

                # Status
                status_str = payment.status.value.capitalize()
                status_item = QTableWidgetItem(status_str)

                if payment.status.value == "completed":
                    status_item.setForeground(QColor("#107C10"))
                elif payment.status.value == "failed":
                    status_item.setForeground(QColor("#D13438"))
                elif payment.status.value == "simulated":
                    status_item.setForeground(QColor("#605E5C"))
                else:
                    status_item.setForeground(QColor("#FF8C00"))

                self.payment_table.setItem(row, 6, status_item)

            # Update summary
            self.summary_label.setText(
                f"Mostrando {len(payments)} de {total_count} | "
                f"Total kg: {total_kg:.2f} | Total MXN: {format_currency(total_mxn, 'MXN')}"
            )
            self.load_more_btn.setVisible(total_count > 200)

        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Error al cargar historial:\n{str(e)}"
            )
        finally:
            close_session()
    
    def _load_more(self):
        """Append next page of 200 payments to the table."""
        try:
            session = get_session()
            self._history_offset += 200
            more_payments = (
                self._build_filtered_query(session)
                .order_by(Payment.timestamp.desc())
                .offset(self._history_offset)
                .limit(200)
                .all()
            )
            current_rows = self.payment_table.rowCount()
            self.payment_table.setRowCount(current_rows + len(more_payments))

            total_count = self._build_filtered_query(session).count()

            for row_idx, payment in enumerate(more_payments):
                row = current_rows + row_idx
                timestamp_str = format_datetime_display(payment.timestamp)
                item = QTableWidgetItem(timestamp_str)
                item.setData(Qt.UserRole, payment.id)
                self.payment_table.setItem(row, 0, item)

                self.payment_table.setItem(row, 1, QTableWidgetItem(payment.producer.name))

                if payment.delivery:
                    self.payment_table.setItem(row, 2, QTableWidgetItem(f"{float(payment.delivery.weight_kg):.2f}"))
                    self.payment_table.setItem(row, 3, QTableWidgetItem(format_currency(float(payment.delivery.price_per_kg), "MXN")))
                else:
                    self.payment_table.setItem(row, 2, QTableWidgetItem("—"))
                    self.payment_table.setItem(row, 3, QTableWidgetItem("—"))

                if payment.amount_mxn:
                    self.payment_table.setItem(row, 4, QTableWidgetItem(format_currency(float(payment.amount_mxn), "MXN")))
                else:
                    self.payment_table.setItem(row, 4, QTableWidgetItem("—"))

                self.payment_table.setItem(row, 5, QTableWidgetItem(f"{float(payment.amount):.6f} {payment.currency}"))

                status_str = payment.status.value.capitalize()
                status_item = QTableWidgetItem(status_str)
                if payment.status.value == "completed":
                    status_item.setForeground(QColor("#107C10"))
                elif payment.status.value == "failed":
                    status_item.setForeground(QColor("#D13438"))
                elif payment.status.value == "simulated":
                    status_item.setForeground(QColor("#605E5C"))
                else:
                    status_item.setForeground(QColor("#FF8C00"))
                self.payment_table.setItem(row, 6, status_item)

            loaded_total = self.payment_table.rowCount()
            self.load_more_btn.setVisible(loaded_total < total_count)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al cargar más pagos:\n{str(e)}")
        finally:
            close_session()

    def show_payment_details(self, index):
        """Show detailed payment information"""
        try:
            row = index.row()
            payment_id = self.payment_table.item(row, 0).data(Qt.UserRole)
            if not payment_id:
                return
            session = get_session()
            payment = session.query(Payment).filter_by(id=payment_id).first()
            if not payment:
                return

            details = (
                f"Detalles del Pago\n"
                f"{'=' * 50}\n\n"
                f"UETR: {payment.uetr}\n"
                f"Hash XRPL: {payment.xrpl_tx_hash}\n"
                f"Fecha: {format_datetime_display(payment.timestamp)}\n\n"
                f"Productor: {payment.producer.name}\n"
                f"Dirección XRPL: {payment.producer.xrpl_address}\n\n"
            )

            if payment.delivery:
                details += (
                    f"Peso: {payment.delivery.weight_kg:.2f} kg\n"
                    f"Precio/kg: {format_currency(payment.delivery.price_per_kg, 'MXN')}\n"
                    f"Total MXN: {format_currency(payment.delivery.total_mxn, 'MXN')}\n\n"
                )

            details += (
                f"Token: {payment.amount:.6f} {payment.currency}\n"
                f"Estado: {payment.status.value.capitalize()}\n"
            )

            if payment.notes:
                details += f"\nNotas: {payment.notes}\n"

            # Add ISO messages info
            if payment.iso_messages:
                details += f"\nMensajes ISO 20022: {len(payment.iso_messages)}\n"
                for msg in payment.iso_messages:
                    details += f"- {msg.message_type.value}\n"

            QMessageBox.information(self, "Detalles del Pago", details)
                
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Error al mostrar detalles:\n{str(e)}"
            )
        finally:
            close_session()
    
    def export_to_excel(self):
        """Export payment history to Excel"""
        try:
            # Get save location
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar Historial",
                f"historial_pagos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # Get data (respects active filters)
            session = get_session()
            payments = self._build_filtered_query(session).order_by(
                Payment.timestamp.desc()
            ).all()
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Historial de Pagos"
            
            # Headers
            headers = [
                "Fecha/Hora", "UETR", "Hash XRPL", "Productor",
                "Peso (kg)", "Precio/kg", "Total MXN",
                "Token", "Cantidad Token", "Estado", "Notas"
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
            
            # Data
            for row, payment in enumerate(payments, 2):
                ws.cell(row=row, column=1, value=format_datetime_display(payment.timestamp))
                ws.cell(row=row, column=2, value=payment.uetr)
                ws.cell(row=row, column=3, value=payment.xrpl_tx_hash)
                ws.cell(row=row, column=4, value=payment.producer.name)
                
                if payment.delivery:
                    ws.cell(row=row, column=5, value=payment.delivery.weight_kg)
                    ws.cell(row=row, column=6, value=payment.delivery.price_per_kg)
                    ws.cell(row=row, column=7, value=payment.delivery.total_mxn)
                
                ws.cell(row=row, column=8, value=payment.currency)
                ws.cell(row=row, column=9, value=payment.amount)
                ws.cell(row=row, column=10, value=payment.status.value)
                ws.cell(row=row, column=11, value=payment.notes or "")
            
            # Save
            wb.save(file_path)
            
            QMessageBox.information(
                self,
                "Exportación Exitosa",
                f"Historial exportado exitosamente a:\n{file_path}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Error al exportar historial:\n{str(e)}"
            )
        finally:
            close_session()
